from __future__ import annotations

import csv
import io
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Callable

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook as OpenPyxlWorkbook
from openpyxl.worksheet.worksheet import Worksheet


SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm", ".csv"}


class ExcelProcessingError(RuntimeError):
    """Raised when a workbook cannot be loaded, processed, or saved."""


@dataclass(frozen=True)
class LoadedSheet:
    """Cached worksheet data used by the UI."""

    name: str
    rows: list[tuple[Any, ...]]
    max_columns: int

    @property
    def row_count(self) -> int:
        return len(self.rows)


class ExcelService:
    """Loads a workbook and runs data-quality checks on sheet 1."""

    def __init__(self) -> None:
        self.input_path: Path | None = None
        self.workbook: OpenPyxlWorkbook | None = None
        self.issues_workbook: OpenPyxlWorkbook | None = None
        self.issue_records_rows: list[tuple[Any, ...]] = []
        self.loaded_sheets: list[LoadedSheet] = []
        self.primary_key_columns: list[str] = []
        self.reviewed = False
        self.last_check_type: str | None = None

    def load(self, file_path: str | Path) -> list[LoadedSheet]:
        path = Path(file_path).expanduser().resolve()
        if not path.exists() or not path.is_file():
            raise ExcelProcessingError(f"File does not exist: {path}")
        if path.suffix.lower() not in SUPPORTED_EXTENSIONS:
            raise ExcelProcessingError(
                "Only .xlsx, .xlsm, and .csv files are supported."
            )

        if path.suffix.lower() == ".csv":
            workbook = self._load_csv_as_workbook(path)
        else:
            try:
                keep_vba = path.suffix.lower() == ".xlsm"
                workbook = load_workbook(
                    filename=path,
                    read_only=False,
                    data_only=False,
                    keep_vba=keep_vba,
                )
            except Exception as exc:  # openpyxl raises several format-specific errors
                raise ExcelProcessingError(f"Could not open the workbook: {exc}") from exc

        if len(workbook.worksheets) < 1:
            workbook.close()
            raise ExcelProcessingError(
                "The workbook must contain at least one worksheet/tab."
            )

        first_sheet = workbook.worksheets[0]
        loaded = self._build_loaded_sheets([first_sheet])
        primary_key_columns = self._extract_primary_key_columns(first_sheet)
        if not primary_key_columns:
            workbook.close()
            raise ExcelProcessingError(
                "Sheet 1 does not have any header values in row 1 to use as a primary key column."
            )

        self.close()
        self.input_path = path
        self.workbook = workbook
        self.issues_workbook = None
        self.issue_records_rows = []
        self.loaded_sheets = loaded
        self.primary_key_columns = primary_key_columns
        self.reviewed = False
        self.last_check_type = None
        return loaded

    def review(
        self,
        primary_key_column: str,
        progress_callback: Callable[[int], None] | None = None,
    ) -> tuple[int, int]:
        workbook = self._require_workbook()
        first_sheet = workbook.worksheets[0]
        primary_key_col, headers = self._resolve_selected_column(first_sheet, primary_key_column)
        key_col_index = primary_key_col - 1
        key_counts: dict[str, int] = defaultdict(int)
        key_row_numbers: dict[str, list[int]] = defaultdict(list)
        key_display_values: dict[str, str] = {}
        total_rows = max(first_sheet.max_row - 1, 0)
        progress_interval = 2000

        # Pass 1: collect duplicate counts/row numbers by key.
        for row_number, row_values in enumerate(
            first_sheet.iter_rows(
                min_row=2,
                max_row=first_sheet.max_row,
                max_col=first_sheet.max_column,
                values_only=True,
            ),
            start=2,
        ):
            key_raw = row_values[key_col_index] if key_col_index < len(row_values) else None
            key_display = self._display_key_value(key_raw)
            key_normalized = self._normalize_token(key_raw)
            key_counts[key_normalized] += 1
            key_row_numbers[key_normalized].append(row_number)
            if key_normalized not in key_display_values:
                key_display_values[key_normalized] = key_display
            if (
                progress_callback is not None
                and total_rows > 0
                and (row_number - 1) % progress_interval == 0
            ):
                progress_callback(min(45, int(((row_number - 1) * 45) / total_rows)))

        summary_rows: list[tuple[Any, ...]] = [
            (
                primary_key_column,
                "Duplicate Count",
                "Row Numbers",
                "Details",
            )
        ]
        issue_rows: list[tuple[Any, ...]] = [
            (
                "Source Row",
                primary_key_column,
                "Duplicate Count",
                *headers,
            )
        ]

        duplicate_keys = [key for key, count in key_counts.items() if count > 1]
        duplicate_key_set = set(duplicate_keys)
        duplicate_records_count = sum(key_counts[key] for key in duplicate_keys)

        sorted_duplicate_keys = sorted(
            duplicate_keys,
            key=lambda item: (
                key_display_values[item] == "(blank)",
                key_display_values[item].casefold(),
            ),
        )

        for key in sorted_duplicate_keys:
            duplicate_count = key_counts[key]
            row_numbers = key_row_numbers[key]
            display_key = key_display_values[key]
            summary_rows.append(
                (
                    display_key,
                    duplicate_count,
                    ", ".join(str(number) for number in row_numbers),
                    f"{duplicate_count} rows share this primary key",
                )
            )

        # Pass 2: only write rows that belong to duplicate keys.
        issue_rows_by_key: dict[str, list[tuple[Any, ...]]] = defaultdict(list)
        for row_number, row_values in enumerate(
            first_sheet.iter_rows(
                min_row=2,
                max_row=first_sheet.max_row,
                max_col=first_sheet.max_column,
                values_only=True,
            ),
            start=2,
        ):
            key_raw = row_values[key_col_index] if key_col_index < len(row_values) else None
            key_normalized = self._normalize_token(key_raw)
            if key_normalized not in duplicate_key_set:
                continue
            issue_rows_by_key[key_normalized].append(
                (
                    row_number,
                    key_display_values[key_normalized],
                    key_counts[key_normalized],
                    *row_values,
                )
            )
            if (
                progress_callback is not None
                and total_rows > 0
                and (row_number - 1) % progress_interval == 0
            ):
                progress_callback(45 + min(50, int(((row_number - 1) * 50) / total_rows)))

        for key in sorted_duplicate_keys:
            issue_rows.extend(
                sorted(
                    issue_rows_by_key.get(key, []),
                    key=lambda row: row[0],  # source row number
                )
            )

        summary_sheet = LoadedSheet(
            name="Duplicate Summary",
            rows=summary_rows,
            max_columns=max(len(row) for row in summary_rows),
        )
        issues_sheet = LoadedSheet(
            name="Issue Records",
            rows=issue_rows,
            max_columns=max(len(row) for row in issue_rows),
        )
        # Reuse already-loaded sheet-1 cache; avoid re-copying large files during review.
        first_loaded_sheet = self.loaded_sheets[0] if self.loaded_sheets else self._build_loaded_sheets([first_sheet])[0]
        self.loaded_sheets = [first_loaded_sheet, summary_sheet, issues_sheet]
        if progress_callback is not None:
            progress_callback(98)
        self.issues_workbook = self._build_issues_workbook(
            summary_title="Duplicate Summary",
            summary_rows=summary_rows,
            issue_rows=issue_rows,
        )
        self.issue_records_rows = issue_rows
        self.reviewed = True
        self.last_check_type = "duplicate"
        if progress_callback is not None:
            progress_callback(100)
        return len(duplicate_keys), duplicate_records_count

    def null_check(
        self,
        primary_key_column: str,
        progress_callback: Callable[[int], None] | None = None,
    ) -> tuple[int, int]:
        workbook = self._require_workbook()
        first_sheet = workbook.worksheets[0]
        primary_key_col, headers = self._resolve_selected_column(first_sheet, primary_key_column)
        key_col_index = primary_key_col - 1
        total_rows = max(first_sheet.max_row - 1, 0)
        progress_interval = 2000

        issue_rows: list[tuple[Any, ...]] = [
            ("Source Row", primary_key_column, "Issue Type", *headers)
        ]
        null_row_numbers: list[int] = []

        for row_number, row_values in enumerate(
            first_sheet.iter_rows(
                min_row=2,
                max_row=first_sheet.max_row,
                max_col=first_sheet.max_column,
                values_only=True,
            ),
            start=2,
        ):
            selected_value = row_values[key_col_index] if key_col_index < len(row_values) else None
            if self._is_null_or_empty(selected_value):
                null_row_numbers.append(row_number)
                issue_rows.append(
                    (
                        row_number,
                        self._display_key_value(selected_value),
                        "Null/Empty value",
                        *row_values,
                    )
                )

            if (
                progress_callback is not None
                and total_rows > 0
                and (row_number - 1) % progress_interval == 0
            ):
                progress_callback(min(95, int(((row_number - 1) * 95) / total_rows)))

        summary_rows: list[tuple[Any, ...]] = [
            (
                primary_key_column,
                "Null/Empty Count",
                "Row Numbers",
                "Details",
            ),
            (
                "(null or empty)",
                len(null_row_numbers),
                ", ".join(str(number) for number in null_row_numbers),
                f"{len(null_row_numbers)} rows have null or empty values in selected column",
            ),
        ]

        summary_sheet = LoadedSheet(
            name="Null Check Summary",
            rows=summary_rows,
            max_columns=max(len(row) for row in summary_rows),
        )
        issues_sheet = LoadedSheet(
            name="Issue Records",
            rows=issue_rows,
            max_columns=max(len(row) for row in issue_rows),
        )
        first_loaded_sheet = self.loaded_sheets[0] if self.loaded_sheets else self._build_loaded_sheets([first_sheet])[0]
        self.loaded_sheets = [first_loaded_sheet, summary_sheet, issues_sheet]
        self.issues_workbook = self._build_issues_workbook(
            summary_title="Null Check Summary",
            summary_rows=summary_rows,
            issue_rows=issue_rows,
        )
        self.issue_records_rows = issue_rows
        self.reviewed = True
        self.last_check_type = "null"
        if progress_callback is not None:
            progress_callback(100)
        return len(null_row_numbers), len(null_row_numbers)

    def _resolve_selected_column(
        self, sheet: Worksheet, selected_column_name: str
    ) -> tuple[int, list[str]]:
        header_map = self._header_index_map(sheet)
        header_key = selected_column_name.strip().casefold()
        selected_column_idx = header_map.get(header_key)
        if selected_column_idx is None:
            raise ExcelProcessingError(
                f"Selected column '{selected_column_name}' was not found in sheet '{sheet.title}'."
            )
        return selected_column_idx, self._display_headers(sheet)

    @staticmethod
    def _header_index_map(sheet: Worksheet) -> dict[str, int]:
        headers: dict[str, int] = {}
        for column in range(1, sheet.max_column + 1):
            header = sheet.cell(row=1, column=column).value
            if header is None:
                continue
            header_key = str(header).strip().lower()
            if header_key:
                headers[header_key] = column
        return headers

    @staticmethod
    def _normalize_token(value: Any) -> str:
        if value is None:
            return ""
        return str(value).strip().casefold()

    @staticmethod
    def _is_null_or_empty(value: Any) -> bool:
        if value is None:
            return True
        return str(value).strip() == ""

    @staticmethod
    def _display_key_value(value: Any) -> str:
        if value is None:
            return "(blank)"
        as_text = str(value).strip()
        return as_text if as_text else "(blank)"

    @staticmethod
    def _display_headers(sheet: Worksheet) -> list[str]:
        headers: list[str] = []
        for column in range(1, sheet.max_column + 1):
            header = sheet.cell(row=1, column=column).value
            if header is None or str(header).strip() == "":
                headers.append(f"Column {get_column_letter(column)}")
            else:
                headers.append(str(header))
        return headers

    @staticmethod
    def _extract_primary_key_columns(sheet: Worksheet) -> list[str]:
        columns: list[str] = []
        seen_normalized: set[str] = set()
        for column in range(1, sheet.max_column + 1):
            header = sheet.cell(row=1, column=column).value
            if header is None:
                continue
            header_name = str(header).strip()
            if not header_name:
                continue
            normalized = header_name.casefold()
            if normalized in seen_normalized:
                continue
            seen_normalized.add(normalized)
            columns.append(header_name)
        return columns

    def save_as(self, output_path: str | Path) -> Path:
        path = Path(output_path).expanduser().resolve()

        if path.suffix.lower() not in SUPPORTED_EXTENSIONS:
            original_suffix = self.input_path.suffix.lower() if self.input_path else ".xlsx"
            if original_suffix not in SUPPORTED_EXTENSIONS:
                original_suffix = ".xlsx"
            path = path.with_suffix(original_suffix)

        path.parent.mkdir(parents=True, exist_ok=True)
        if path.suffix.lower() == ".csv":
            rows = self._require_issue_records_rows()
            self._save_issue_rows_as_csv(path, rows)
            return path

        workbook = self._require_issues_workbook()
        try:
            workbook.save(path)
        except Exception as exc:
            raise ExcelProcessingError(f"Could not save the workbook: {exc}") from exc
        return path

    def default_output_path(self) -> Path:
        if self.input_path is None:
            raise ExcelProcessingError("Load a workbook first.")
        suffix = "issues"
        if self.last_check_type == "duplicate":
            suffix = "duplicate_issues"
        elif self.last_check_type == "null":
            suffix = "null_issues"
        return self.input_path.with_name(
            f"{self.input_path.stem}_{suffix}{self.input_path.suffix}"
        )

    def close(self) -> None:
        if self.workbook is not None:
            try:
                self.workbook.close()
            except Exception:
                pass
        if self.issues_workbook is not None:
            try:
                self.issues_workbook.close()
            except Exception:
                pass
        self.workbook = None
        self.issues_workbook = None
        self.issue_records_rows = []
        self.loaded_sheets = []
        self.primary_key_columns = []
        self.input_path = None
        self.reviewed = False
        self.last_check_type = None

    def _require_workbook(self) -> OpenPyxlWorkbook:
        if self.workbook is None:
            raise ExcelProcessingError("Load an Excel workbook first.")
        return self.workbook

    def _require_issues_workbook(self) -> OpenPyxlWorkbook:
        if self.issues_workbook is None or not self.reviewed:
            raise ExcelProcessingError("Review the workbook first to generate issue records.")
        return self.issues_workbook

    def _require_issue_records_rows(self) -> list[tuple[Any, ...]]:
        if not self.reviewed or not self.issue_records_rows:
            raise ExcelProcessingError("Review the workbook first to generate issue records.")
        return self.issue_records_rows

    @staticmethod
    def _build_loaded_sheets(sheets: list[Worksheet]) -> list[LoadedSheet]:
        loaded: list[LoadedSheet] = []
        for worksheet in sheets:
            rows = [tuple(row) for row in worksheet.iter_rows(values_only=True)]
            max_columns = max((len(row) for row in rows), default=0)
            loaded.append(
                LoadedSheet(
                    name=worksheet.title,
                    rows=rows,
                    max_columns=max_columns,
                )
            )
        return loaded

    @staticmethod
    def _build_issues_workbook(
        summary_title: str,
        summary_rows: list[tuple[Any, ...]],
        issue_rows: list[tuple[Any, ...]],
    ) -> OpenPyxlWorkbook:
        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = summary_title
        for row in summary_rows:
            summary_sheet.append(list(row))

        issues_sheet = workbook.create_sheet("Issue Records")
        for row in issue_rows:
            issues_sheet.append(list(row))
        return workbook

    @staticmethod
    def _load_csv_as_workbook(path: Path) -> OpenPyxlWorkbook:
        workbook = Workbook()
        sheet = workbook.active
        clean_title = ExcelService._sanitize_sheet_title(path.stem) or "Sheet1"
        sheet.title = clean_title
        try:
            csv_text = ExcelService._decode_csv_bytes(path)
            reader = csv.reader(io.StringIO(csv_text))
            for row in reader:
                sheet.append(row)
        except Exception as exc:
            workbook.close()
            raise ExcelProcessingError(f"Could not open CSV file: {exc}") from exc
        return workbook

    @staticmethod
    def _decode_csv_bytes(path: Path) -> str:
        raw = path.read_bytes()
        if raw.startswith(b"\xef\xbb\xbf"):
            raw = raw[3:]
        try:
            return raw.decode("utf-8")
        except UnicodeDecodeError:
            # Fallback for files saved from legacy Excel/system locales.
            return raw.decode("cp1252")

    @staticmethod
    def _save_issue_rows_as_csv(path: Path, rows: list[tuple[Any, ...]]) -> None:
        try:
            with path.open("w", newline="", encoding="utf-8") as handle:
                writer = csv.writer(handle)
                for row in rows:
                    writer.writerow(["" if value is None else value for value in row])
        except Exception as exc:
            raise ExcelProcessingError(f"Could not save CSV file: {exc}") from exc

    @staticmethod
    def _sanitize_sheet_title(title: str) -> str:
        invalid_chars = {":", "\\", "/", "?", "*", "[", "]"}
        cleaned = "".join("_" if char in invalid_chars else char for char in title).strip()
        if not cleaned:
            return "Sheet1"
        return cleaned[:31]


def create_sample_workbook(output_path: str | Path) -> Path:
    """Create a small workbook that includes duplicate values for testing."""
    path = Path(output_path).expanduser().resolve()
    workbook = Workbook()
    first_sheet = workbook.active
    first_sheet.title = "Data"
    first_sheet.append(["Record ID", "Name", "Country", "Amount"])
    first_sheet.append([1001, "Alice", "US", 50])
    first_sheet.append([1002, "Bob", "US", 75])
    first_sheet.append([1001, "Alice Duplicate", "US", 80])
    first_sheet.append([1003, "Charlie", "CA", 20])
    first_sheet.append([1002, "Bob Duplicate", "UK", 95])
    first_sheet.append([None, "Missing PK #1", "FR", 10])
    first_sheet.append([None, "Missing PK #2", "DE", 11])

    workbook.create_sheet("Other Sheet")

    workbook.save(path)
    workbook.close()
    return path
