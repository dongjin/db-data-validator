import csv
from pathlib import Path

from openpyxl import load_workbook

from db_data_validator.excel_service import ExcelService, create_sample_workbook


def test_review_duplicates_and_save_excel_issues(tmp_path: Path) -> None:
    input_path = create_sample_workbook(tmp_path / "input.xlsx")
    service = ExcelService()

    loaded = service.load(input_path)
    assert [sheet.name for sheet in loaded] == ["Data"]
    assert loaded[0].row_count == 8
    assert service.primary_key_columns == ["Record ID", "Name", "Country", "Amount"]

    duplicate_keys, duplicate_rows = service.review("Record ID")
    assert duplicate_keys == 3
    assert duplicate_rows == 6
    assert [sheet.name for sheet in service.loaded_sheets] == [
        "Data",
        "Duplicate Summary",
        "Issue Records",
    ]

    output_path = service.save_as(tmp_path / "output.xlsx")
    service.close()

    workbook = load_workbook(output_path)
    assert workbook.sheetnames == ["Duplicate Summary", "Issue Records"]
    summary = workbook["Duplicate Summary"]
    assert summary["A2"].value == "(blank)"
    assert summary["B2"].value == 2
    issues = workbook["Issue Records"]
    assert issues["A1"].value == "Source Row"
    assert issues["B1"].value == "Record ID"
    assert issues.max_row == 7  # header + 6 issue rows
    workbook.close()


def test_load_csv_review_and_save_csv_issues(tmp_path: Path) -> None:
    input_path = tmp_path / "input.csv"
    input_path.write_text(
        "id,name,region\n"
        "1001,Alice,US\n"
        "1002,Bob,US\n"
        "1001,Alice Again,CA\n"
        ",Missing 1,DE\n"
        ",Missing 2,FR\n",
        encoding="utf-8",
    )
    service = ExcelService()

    loaded = service.load(input_path)
    assert loaded[0].name == "input"
    assert service.primary_key_columns == ["id", "name", "region"]

    duplicate_keys, duplicate_rows = service.review("id")
    assert duplicate_keys == 2
    assert duplicate_rows == 4

    output_path = service.save_as(tmp_path / "issues.csv")
    service.close()

    with output_path.open("r", newline="", encoding="utf-8") as handle:
        rows = list(csv.reader(handle))

    assert rows[0][:3] == ["Source Row", "id", "Duplicate Count"]
    assert len(rows) == 5  # header + 4 issue rows
